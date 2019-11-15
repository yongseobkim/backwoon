from django.shortcuts import render, redirect

# Create your views here.

from django.contrib.auth.decorators import login_required

import boto3
import io
import pandas as pd
import xlrd as rd
import openpyxl as xl
from django.conf import settings
from config import asset_storage
from input.models import Files
from .models import Data
from django.db.models import Max, Min, Sum
import datetime

@login_required

def loading(request):


    aws_access_key_id = settings.AWS_ACCESS_KEY_ID
    aws_secret_access_key = settings.AWS_SECRET_ACCESS_KEY
    bucket = settings.AWS_STORAGE_BUCKET_NAME

    file = Files.objects.all()[0]
    key1 = file.file.name
    key2 = asset_storage.FileStorage.location
    key = key2+'/'+key1

    s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key)
    obj = s3.get_object(Bucket=bucket, Key=key)

    data = obj['Body'].read()
    #df = pd.read_excel(io.BytesIO(data), encoding='utf-8')
    df = xl.load_workbook(io.BytesIO(data))
    sheet = df['Sheet1']

    cp = '2100-01-01'
    cp = datetime.datetime.strptime(cp, "%Y-%m-%d").date()
    for i in range(2, sheet.max_row+1):
        a_rdate = sheet.cell(i, 2).value
        a_rdate = datetime.datetime.strptime(a_rdate, "%Y-%m-%d").date()
        if cp >= a_rdate: cp = a_rdate
    #print("@@@@@@@@@@",cp)
    #datas = Data.objects.aggregate(a_rdate = Min('a_rdate'))
    datas = Data.objects.filter(a_rdate__gte=cp)
    datas.delete()

    #a = sheet.cell(4,2).value
    #a = datetime.datetime.strptime(a, "%Y-%m-%d").date()
    #print(type(a),a)
    for i in range(2, sheet.max_row +1):
        a_rdate = sheet.cell(i,2).value
        a_rdate = datetime.datetime.strptime(a_rdate, "%Y-%m-%d").date()
        car_num2 = sheet.cell(i, 3).value
        car_num3 = sheet.cell(i, 4).value
        car_num1 = sheet.cell(i, 5).value
        fwc_sdate = sheet.cell(i, 6).value
        try:
            fwc_sdate = datetime.datetime.strptime(fwc_sdate, "%Y-%m-%d").date()
        except:
            fwc_sdate = None
        car_type = sheet.cell(i, 7).value
        car_year = sheet.cell(i, 8).value
        car_model = sheet.cell(i, 9).value
        car_cor = sheet.cell(i,10).value
        a_limm = int(sheet.cell(i, 11).value)
        cus_name = sheet.cell(i, 12).value
        cus_add = sheet.cell(i, 13).value
        cus_add2 = sheet.cell(i, 14).value
        a_type = sheet.cell(i, 15).value
        car_price = int(sheet.cell(i, 16).value)
        car_dc = int(sheet.cell(i, 17).value)
        fwc_stdprice = int(sheet.cell(i, 18).value)
        fwc_limm = int(sheet.cell(i, 19).value)
        fwc_pr_code = sheet.cell(i, 20).value
        fwc_pt_dept = sheet.cell(i, 21).value
        fwc_3pt = sheet.cell(i, 22).value
        a_mgr = sheet.cell(i, 23).value
        fwc_pt_dept_mgr = sheet.cell(i, 24).value
        car_ms =sheet.cell(i, 25).value
        text = sheet.cell(i, 31).value
        a_spot = sheet.cell(i,32).value
        a_move = sheet.cell(i, 33).value
        car_ap = sheet.cell(i, 35).value
        car_ap2 = sheet.cell(i, 36).value
        car_acd = sheet.cell(i, 37).value
        car_sdate = sheet.cell(i, 38).value
        try :
            car_sdate = datetime.datetime.strptime(car_sdate, "%Y-%m-%d").date()
        except:
            car_sdate = None
        try :
            data = Data.objects.get(a_rdate=a_rdate, car_num2=car_num2)
            data.a_rdate = a_rdate
            data.car_num2 = car_num2
            data.car_num3 = car_num3
            data.car_num1 = car_num1
            data.fwc_sdate = fwc_sdate
            data.car_type = car_type
            data.car_year = car_year
            data.car_model = car_model
            data.car_cor = car_cor
            data.a_limm = a_limm
            data.cus_name = cus_name
            data.cus_add = cus_add
            data.cus_add2 = cus_add2
            data.a_type = a_type
            data.car_price = car_price
            data.car_dc = car_dc
            data.fwc_stdprice =fwc_stdprice
            data.fwc_limm = fwc_limm
            data.fwc_pr_code = fwc_pr_code
            data.fwc_pt_dept = fwc_pt_dept
            data.fwc_3pt = fwc_3pt
            data.a_mgr = a_mgr
            data.fwc_pt_dept_mgr = fwc_pt_dept_mgr
            data.car_ms = car_ms
            data.text = text
            data.a_spot = a_spot
            data.a_move = a_move
            data.car_ap = car_ap
            data.car_ap2 = car_ap2
            data.car_acd = car_acd
            data.car_sdate = car_sdate
            data.save()
        except :
            data = Data(
                a_rdate = a_rdate,
                car_num2 = car_num2,
                car_num3 = car_num3,
                car_num1 = car_num1,
                fwc_sdate = fwc_sdate,
                car_type = car_type,
                car_year = car_year,
                car_model = car_model,
                car_cor = car_cor,
                a_limm = a_limm,
                cus_name = cus_name,
                cus_add = cus_add,
                cus_add2 = cus_add2,
                a_type = a_type,
                car_price = car_price,
                car_dc = car_dc,
                fwc_stdprice = fwc_stdprice,
                fwc_limm = fwc_limm,
                fwc_pr_code = fwc_pr_code,
                fwc_pt_dept = fwc_pt_dept,
                fwc_3pt = fwc_3pt,
                a_mgr = a_mgr,
                fwc_pt_dept_mgr = fwc_pt_dept_mgr,
                car_ms = car_ms,
                text = text,
                a_spot = a_spot,
                a_move = a_move,
                car_ap = car_ap,
                car_ap2 = car_ap2,
                car_acd = car_acd,
                car_sdate = car_sdate
            )
            data.save()

        #print(a_rdate, car_num2)
        #Data.a_rdate = sheet.cell(i, 2).value
        #Data.save()

    #cell = sheet.cell(3, 4).value
    #print("@@@@@@@@@@ 값은? ", sheet.max_row)

    return redirect('/data')
